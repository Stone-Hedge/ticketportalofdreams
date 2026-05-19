import csv
from datetime import datetime
from openpyxl import load_workbook, Workbook
from .models import Person, Seat, Fixture, TicketAllocation, SensitiveDetail

def import_workbook(path):
    wb = load_workbook(path)
    report = {'people':0,'seats':0,'fixtures':0,'allocations':0,'sensitive':0}
    seats_sheet = wb['SEATS'] if 'SEATS' in wb.sheetnames else wb.active
    headers = [c.value for c in seats_sheet[1]]
    seat_cols = []
    for idx, h in enumerate(headers[6:], start=7):
        if h:
            seat, created = Seat.objects.get_or_create(label=str(h), defaults={'block':'?', 'row':'?', 'seat_number':str(h).split()[-1]})
            report['seats'] += int(created); seat_cols.append((idx, seat))
    for row in seats_sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]: continue
        dt = row[0] if hasattr(row[0],'year') else datetime.strptime(str(row[0]), '%Y-%m-%d').date()
        fixture, c = Fixture.objects.get_or_create(date=dt, opponent=str(row[1] or 'TBD'), defaults={'competition':row[2] or '', 'category':row[4] or ''})
        report['fixtures'] += int(c)
        for col, seat in seat_cols:
            val = row[col-1] if len(row)>=col else None
            alloc, ac = TicketAllocation.objects.get_or_create(fixture=fixture, seat=seat)
            report['allocations'] += int(ac)
            if val:
                p,_=Person.objects.get_or_create(name=str(val), defaults={'short_name':str(val).split()[0]})
                alloc.assigned_to=p; alloc.payment_status='unpaid'; alloc.save()
    if 'DETAILS' in wb.sheetnames:
        for r in wb['DETAILS'].iter_rows(min_row=2, values_only=True):
            if not r or not r[0]: continue
            p,_=Person.objects.get_or_create(name=str(r[0]), defaults={'short_name':str(r[0]).split()[0]})
            _,cr=SensitiveDetail.objects.get_or_create(person=p, defaults={'arsenal_membership_number':str(r[1] or ''), 'bank_details':str(r[2] or ''), 'notes':str(r[3] or '')})
            report['sensitive']+=int(cr)
    report['people']=Person.objects.count()
    return report

def export_xlsx(path):
    wb = Workbook(); ws = wb.active; ws.title='SEATS'
    seats = list(Seat.objects.filter(active=True))
    ws.append(['date','opponent','competition','kick_off','category','face_value'] + [s.label for s in seats])
    for f in Fixture.objects.all().order_by('date'):
        base=[f.date.isoformat(),f.opponent,f.competition,f.kick_off.isoformat() if f.kick_off else '',f.category,f.face_value or '']
        names=[]
        for s in seats:
            a=TicketAllocation.objects.filter(fixture=f, seat=s).first()
            names.append(a.assigned_to.name if a and a.assigned_to else '')
        ws.append(base+names)
    wb.save(path)

def export_csv(path):
    seats=list(Seat.objects.filter(active=True))
    with open(path,'w',newline='') as f:
        w=csv.writer(f); w.writerow(['date','opponent']+[s.label for s in seats])
        for fi in Fixture.objects.all().order_by('date'):
            row=[fi.date.isoformat(), fi.opponent]
            for s in seats:
                a=TicketAllocation.objects.filter(fixture=fi, seat=s).first(); row.append(a.assigned_to.name if a and a.assigned_to else '')
            w.writerow(row)
